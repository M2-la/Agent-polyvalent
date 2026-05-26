"""
bi_tool.py — Business Intelligence & Reporting avancé
Bibliothèques utilisées selon le contexte :
  - Plotly      ★ Dashboards interactifs, web, partage
  - Matplotlib  ★ Graphiques statiques, publication, PDF
  - Seaborn     ★ Analyses statistiques, distributions, corrélations
  - DuckDB      ★ SQL analytique ultra-rapide sur les données

Formats d'export :
  - HTML   Dashboard interactif (Plotly CDN, zéro dépendance)
  - PDF    Rapport professionnel (ReportLab mise en page avancée)
  - PNG    Image statique haute résolution (Matplotlib/Seaborn)
  - XLSX   Export Excel avec mise en forme

Planification : APScheduler (cron)
"""

import os
import json
import asyncio
import hashlib
from pathlib import Path
from datetime import datetime
from typing import Any
from dotenv import load_dotenv

load_dotenv()

REPORTS_DIR = os.getenv("REPORTS_DIR", "reports")
BI_API_URL   = os.getenv("BI_API_URL", "")
BI_API_KEY   = os.getenv("BI_API_KEY", "")

# Types de graphiques → librairie recommandée
CHART_LIBRARY = {
    # Interactif / web
    "bar": "plotly", "line": "plotly", "pie": "plotly",
    "scatter": "plotly", "area": "plotly", "funnel": "plotly",
    "treemap": "plotly", "sunburst": "plotly", "gauge": "plotly",
    "waterfall": "plotly", "candlestick": "plotly",
    # Stats / distribution
    "histogram": "seaborn", "boxplot": "seaborn", "violin": "seaborn",
    "heatmap": "seaborn", "pairplot": "seaborn", "kde": "seaborn",
    "regression": "seaborn",
    # Publication / PDF
    "static_bar": "matplotlib", "static_line": "matplotlib",
    "static_pie": "matplotlib", "dashboard": "plotly",
}


class BITool:
    """Business Intelligence : rapports depuis APIs, fichiers, bases de données."""

    def __init__(self):
        Path(REPORTS_DIR).mkdir(parents=True, exist_ok=True)
        self._scheduler = None

    # ─────────────────────────────────────────────
    # Fetch API → DataFrame
    # ─────────────────────────────────────────────

    async def fetch_api_data(
        self,
        url: str,
        method: str = "GET",
        headers: dict = None,
        params: dict = None,
        body: dict = None,
        api_key: str = None,
        auth_type: str = "bearer",
    ) -> dict:
        """Récupérer des données depuis une API REST."""
        try:
            import aiohttp

            req_headers = {"Accept": "application/json", "Content-Type": "application/json"}

            # Headers depuis .env
            env_h = os.getenv("BI_API_HEADERS", "")
            if env_h:
                try:
                    req_headers.update(json.loads(env_h))
                except Exception:
                    pass
            if headers:
                req_headers.update(headers)

            # Auth
            key = api_key or BI_API_KEY
            if key:
                if auth_type == "bearer":
                    req_headers["Authorization"] = f"Bearer {key}"
                elif auth_type == "api_key":
                    req_headers["X-API-Key"] = key
                elif auth_type == "basic":
                    import base64
                    req_headers["Authorization"] = f"Basic {base64.b64encode(key.encode()).decode()}"

            async with aiohttp.ClientSession() as session:
                async with session.request(
                    method.upper(), url,
                    headers=req_headers, params=params,
                    json=body, timeout=aiohttp.ClientTimeout(total=30), ssl=False,
                ) as resp:
                    ct = resp.headers.get("Content-Type", "")
                    if "json" in ct:
                        data = await resp.json()
                    else:
                        text = await resp.text()
                        try:
                            data = json.loads(text)
                        except Exception:
                            data = {"raw": text}

                    if resp.status >= 400:
                        return {"success": False, "error": f"HTTP {resp.status}", "data": data}
                    return {"success": True, "data": data, "status_code": resp.status}

        except Exception as e:
            return {"success": False, "error": str(e), "data": None}

    def normalize_data(self, raw: Any, path_hint: str = None) -> "pd.DataFrame | None":
        """Convertir n'importe quelle source en DataFrame."""
        try:
            import pandas as pd
            import io

            if isinstance(raw, pd.DataFrame):
                return raw
            if isinstance(raw, list) and raw:
                return pd.DataFrame(raw)
            if isinstance(raw, dict):
                # Chercher une liste imbriquée
                for v in raw.values():
                    if isinstance(v, list) and v:
                        try:
                            return pd.DataFrame(v)
                        except Exception:
                            pass
                # Sinon : dict plat
                return pd.DataFrame([raw])
            if isinstance(raw, str):
                # Essayer CSV
                try:
                    return pd.read_csv(io.StringIO(raw))
                except Exception:
                    pass
            return None
        except Exception:
            return None

    # ─────────────────────────────────────────────
    # ★ Plotly — graphiques interactifs
    # ─────────────────────────────────────────────

    async def plotly_chart(
        self, data: Any, chart_type: str = "bar",
        x: str = None, y: str = None, color: str = None,
        title: str = "Graphique", height: int = 450,
    ) -> str:
        """Graphique Plotly interactif → HTML."""
        return await asyncio.to_thread(
            self._plotly_sync, data, chart_type, x, y, color, title, height
        )

    def _plotly_sync(self, data, chart_type, x, y, color, title, height) -> str:
        import plotly.express as px
        import plotly.graph_objects as go
        import pandas as pd

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        if df is None or df.empty:
            return None

        num = df.select_dtypes(include="number").columns.tolist()
        cat = df.select_dtypes(include="object").columns.tolist()
        x = x or (cat[0] if cat else (df.columns[0] if len(df.columns) > 0 else None))
        y = y or (num[0] if num else (df.columns[1] if len(df.columns) > 1 else None))

        fig_args = dict(template="plotly_white", title=title, height=height)
        try:
            ct = chart_type.lower()
            if ct == "bar":
                fig = px.bar(df, x=x, y=y, color=color, text_auto=True, **fig_args)
            elif ct == "line":
                fig = px.line(df, x=x, y=y, color=color, markers=True, **fig_args)
            elif ct == "pie":
                fig = px.pie(df, names=x, values=y, hole=0.3, **fig_args)
            elif ct == "scatter":
                fig = px.scatter(df, x=x, y=y, color=color, **fig_args)
            elif ct == "area":
                fig = px.area(df, x=x, y=y, color=color, **fig_args)
            elif ct == "funnel":
                fig = px.funnel(df, x=y, y=x, **fig_args)
            elif ct == "treemap":
                path_cols = cat[:3] if cat else [x]
                fig = px.treemap(df, path=path_cols, values=y, **fig_args)
            elif ct == "waterfall":
                fig = go.Figure(go.Waterfall(x=df[x], y=df[y], name=title))
                fig.update_layout(**{k: v for k, v in fig_args.items() if k != "template"})
            elif ct == "candlestick":
                # Attend des colonnes open/high/low/close
                fig = go.Figure(go.Candlestick(
                    x=df[x],
                    open=df.get("open", df[y]), high=df.get("high", df[y]),
                    low=df.get("low", df[y]), close=df.get("close", df[y])
                ))
                fig.update_layout(title=title, height=height)
            elif ct == "histogram":
                import plotly.express as px
                fig = px.histogram(df, x=y or x, color=color, nbins=30, **fig_args)
            elif ct == "gauge":
                val = df[y].mean() if y in df else 0
                fig = go.Figure(go.Indicator(
                    mode="gauge+number", value=val,
                    title={"text": title},
                    gauge={"axis": {"range": [0, df[y].max() * 1.2 if y in df else 100]},
                           "bar": {"color": "#1a237e"}}
                ))
            else:
                fig = px.bar(df, x=x, y=y, **fig_args)

            fig.update_layout(
                font=dict(family="Segoe UI, sans-serif", size=13),
                margin=dict(l=40, r=30, t=60, b=40),
            )

            ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            out = os.path.join(REPORTS_DIR, f"chart_{ct}_{ts}.html")
            fig.write_html(out, include_plotlyjs="cdn")
            return out

        except Exception as e:
            print(f"[BITool] plotly_chart error: {e}")
            return None

    # ─────────────────────────────────────────────
    # ★ Seaborn — graphiques statistiques
    # ─────────────────────────────────────────────

    async def seaborn_chart(
        self, data: Any, chart_type: str = "heatmap",
        x: str = None, y: str = None, hue: str = None,
        title: str = "Analyse statistique",
    ) -> str:
        """Graphique Seaborn statistique → PNG."""
        return await asyncio.to_thread(
            self._seaborn_sync, data, chart_type, x, y, hue, title
        )

    def _seaborn_sync(self, data, chart_type, x, y, hue, title) -> str:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import seaborn as sns
        import pandas as pd

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        if df is None or df.empty:
            return None

        num = df.select_dtypes(include="number").columns.tolist()
        cat = df.select_dtypes(include="object").columns.tolist()
        x = x or (cat[0] if cat else (num[0] if num else df.columns[0]))
        y = y or (num[0] if num else df.columns[-1])

        sns.set_theme(style="whitegrid", palette="muted")
        ct = chart_type.lower()

        try:
            if ct == "heatmap":
                fig, ax = plt.subplots(figsize=(10, 8))
                corr = df[num].corr() if len(num) >= 2 else df.corr()
                sns.heatmap(corr, annot=True, fmt=".2f", cmap="coolwarm",
                            vmin=-1, vmax=1, ax=ax)
                ax.set_title(title)

            elif ct == "boxplot":
                fig, ax = plt.subplots(figsize=(12, 6))
                if y and x:
                    sns.boxplot(data=df, x=x, y=y, hue=hue, ax=ax)
                else:
                    sns.boxplot(data=df[num[:6]], ax=ax)
                ax.set_title(title)
                plt.xticks(rotation=30)

            elif ct == "violin":
                fig, ax = plt.subplots(figsize=(12, 6))
                sns.violinplot(data=df, x=x, y=y, hue=hue, ax=ax, inner="box")
                ax.set_title(title)
                plt.xticks(rotation=30)

            elif ct == "histogram":
                cols = [y] if y else num[:4]
                fig, axes = plt.subplots(1, len(cols), figsize=(5 * len(cols), 5))
                if len(cols) == 1:
                    axes = [axes]
                for ax, col in zip(axes, cols):
                    sns.histplot(df[col].dropna(), kde=True, ax=ax)
                    ax.set_title(col)
                fig.suptitle(title)

            elif ct == "pairplot":
                cols = num[:5]  # Limiter
                g = sns.pairplot(df[cols].dropna(), diag_kind="kde", plot_kws={"alpha": 0.5})
                g.fig.suptitle(title, y=1.02)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out = os.path.join(REPORTS_DIR, f"chart_pairplot_{ts}.png")
                g.savefig(out, dpi=150, bbox_inches="tight")
                plt.close()
                return out

            elif ct == "kde":
                fig, ax = plt.subplots(figsize=(10, 6))
                for col in num[:5]:
                    sns.kdeplot(df[col].dropna(), label=col, ax=ax, fill=True, alpha=0.3)
                ax.set_title(title)
                ax.legend()

            elif ct == "regression":
                fig, ax = plt.subplots(figsize=(10, 6))
                sns.regplot(data=df, x=x, y=y, ax=ax, scatter_kws={"alpha": 0.5})
                ax.set_title(f"Régression : {x} → {y}")

            else:
                fig, ax = plt.subplots(figsize=(10, 6))
                sns.heatmap(df[num].corr(), annot=True, fmt=".2f", cmap="coolwarm", ax=ax)
                ax.set_title(title)

            plt.tight_layout()
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(REPORTS_DIR, f"chart_{ct}_{ts}.png")
            plt.savefig(out, dpi=150, bbox_inches="tight")
            plt.close()
            return out

        except Exception as e:
            print(f"[BITool] seaborn error: {e}")
            try:
                plt.close()
            except Exception:
                pass
            return None

    # ─────────────────────────────────────────────
    # ★ Matplotlib — graphiques statiques publication
    # ─────────────────────────────────────────────

    async def matplotlib_chart(
        self, data: Any, chart_type: str = "bar",
        x: str = None, y: str = None, title: str = "Graphique",
    ) -> str:
        """Graphique Matplotlib statique → PNG haute résolution."""
        return await asyncio.to_thread(self._mpl_sync, data, chart_type, x, y, title)

    def _mpl_sync(self, data, chart_type, x, y, title) -> str:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import pandas as pd

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        if df is None or df.empty:
            return None

        num = df.select_dtypes(include="number").columns.tolist()
        cat = df.select_dtypes(include="object").columns.tolist()
        x = x or (cat[0] if cat else df.columns[0])
        y = y or (num[0] if num else df.columns[-1])

        fig, ax = plt.subplots(figsize=(12, 6))
        ct = chart_type.lower()

        try:
            if ct in ("bar", "static_bar") and x in df.columns and y in df.columns:
                if df[x].nunique() <= 20:
                    grp = df.groupby(x)[y].sum().sort_values(ascending=False)
                    bars = ax.bar(grp.index, grp.values, color="#1a237e", edgecolor="white")
                    ax.bar_label(bars, fmt="%.1f", padding=3, fontsize=9)
                    plt.xticks(rotation=45, ha="right")
                else:
                    ax.bar(range(len(df)), df[y], color="#1a237e")
                    ax.set_xlabel(x)

            elif ct in ("line", "static_line") and x in df.columns and y in df.columns:
                ax.plot(df[x], df[y], marker="o", color="#1a237e", linewidth=2, markersize=5)
                plt.xticks(rotation=45, ha="right")

            elif ct in ("pie", "static_pie"):
                data_pie = df.groupby(x)[y].sum().sort_values(ascending=False).head(10)
                ax.pie(data_pie.values, labels=data_pie.index,
                       autopct="%1.1f%%", startangle=140,
                       colors=plt.cm.Set3.colors)

            else:
                ax.bar(range(len(df)), df[num[0]] if num else df.iloc[:, 0], color="#1a237e")

            ax.set_title(title, fontsize=14, fontweight="bold", pad=15)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.grid(axis="y", alpha=0.3)
            plt.tight_layout()

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = os.path.join(REPORTS_DIR, f"chart_mpl_{ct}_{ts}.png")
            plt.savefig(out, dpi=200, bbox_inches="tight")
            plt.close()
            return out

        except Exception as e:
            print(f"[BITool] matplotlib error: {e}")
            plt.close()
            return None

    # ─────────────────────────────────────────────
    # ★ Dashboard HTML complet (multi-graphiques)
    # ─────────────────────────────────────────────

    async def generate_dashboard(
        self,
        data: Any,
        title: str = "Dashboard BI",
        chart_configs: list[dict] = None,
        api_url: str = "",
        theme: str = "blue",  # blue | green | dark
    ) -> str:
        """Dashboard HTML avec KPIs + multiple graphiques Plotly."""
        return await asyncio.to_thread(
            self._dashboard_sync, data, title, chart_configs, api_url, theme
        )

    def _dashboard_sync(self, data, title, chart_configs, api_url, theme) -> str:
        import plotly.express as px
        import plotly.io as pio
        import plotly.graph_objects as go
        import pandas as pd

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        if df is None or df.empty:
            return None

        num = df.select_dtypes(include="number").columns.tolist()
        cat = df.select_dtypes(include="object").columns.tolist()
        dt_cols = df.select_dtypes(include=["datetime64"]).columns.tolist()

        # Config auto si non fournie
        if not chart_configs:
            chart_configs = []
            if cat and num:
                chart_configs.append({"type": "bar", "x": cat[0], "y": num[0], "title": f"{num[0]} par {cat[0]}"})
            if len(num) >= 2:
                chart_configs.append({"type": "scatter", "x": num[0], "y": num[1], "title": f"{num[0]} vs {num[1]}"})
            if cat and num:
                chart_configs.append({"type": "pie", "x": cat[0], "y": num[0], "title": f"Répartition {cat[0]}"})
            if dt_cols and num:
                chart_configs.append({"type": "line", "x": dt_cols[0], "y": num[0], "title": f"Évolution {num[0]}"})
            if len(num) >= 2:
                chart_configs.append({"type": "heatmap_corr", "title": "Corrélations"})
            if num:
                chart_configs.append({"type": "histogram", "x": num[0], "title": f"Distribution {num[0]}"})

        # Thèmes
        themes = {
            "blue":  {"primary": "#1a237e", "secondary": "#0d47a1", "accent": "#e8eaf6", "text": "#fff"},
            "green": {"primary": "#1b5e20", "secondary": "#2e7d32", "accent": "#e8f5e9", "text": "#fff"},
            "dark":  {"primary": "#212121", "secondary": "#424242", "accent": "#f5f5f5", "text": "#fff"},
        }
        t = themes.get(theme, themes["blue"])

        # Générer chaque graphique
        chart_htmls = []
        for cfg in chart_configs[:8]:
            ct = cfg.get("type", "bar")
            cx = cfg.get("x")
            cy = cfg.get("y")
            ctitle = cfg.get("title", ct)
            c_color = cfg.get("color")

            try:
                if ct == "heatmap_corr":
                    if len(num) < 2:
                        continue
                    corr = df[num[:8]].corr()
                    fig = px.imshow(corr, title=ctitle, color_continuous_scale="RdBu_r",
                                    zmin=-1, zmax=1, template="plotly_white", height=350)
                elif ct == "bar" and cx and cy and cx in df.columns and cy in df.columns:
                    if df[cx].nunique() <= 25:
                        grp = df.groupby(cx)[cy].sum().sort_values(ascending=False).reset_index()
                        fig = px.bar(grp, x=cx, y=cy, title=ctitle, template="plotly_white",
                                     height=350, text_auto=True, color_discrete_sequence=[t["primary"]])
                    else:
                        fig = px.histogram(df, x=cy, title=ctitle, template="plotly_white", height=350)
                elif ct == "line" and cx and cy and cx in df.columns and cy in df.columns:
                    fig = px.line(df.sort_values(cx), x=cx, y=cy, title=ctitle,
                                  template="plotly_white", height=350, markers=True,
                                  color_discrete_sequence=[t["primary"]])
                elif ct == "pie" and cx and cy and cx in df.columns and cy in df.columns:
                    grp = df.groupby(cx)[cy].sum().nlargest(10).reset_index()
                    fig = px.pie(grp, names=cx, values=cy, title=ctitle,
                                 template="plotly_white", height=350, hole=0.35)
                elif ct == "scatter" and cx and cy and cx in df.columns and cy in df.columns:
                    fig = px.scatter(df, x=cx, y=cy, title=ctitle, color=c_color,
                                     template="plotly_white", height=350, opacity=0.7)
                elif ct == "histogram":
                    col = cy or cx or (num[0] if num else None)
                    if not col:
                        continue
                    fig = px.histogram(df, x=col, title=ctitle, template="plotly_white",
                                       height=350, nbins=30, color_discrete_sequence=[t["primary"]])
                elif ct == "area" and cx and cy and cx in df.columns and cy in df.columns:
                    fig = px.area(df.sort_values(cx), x=cx, y=cy, title=ctitle,
                                  template="plotly_white", height=350)
                else:
                    continue

                fig.update_layout(
                    font=dict(family="Segoe UI, sans-serif", size=12),
                    margin=dict(l=30, r=20, t=50, b=30),
                    plot_bgcolor="white",
                )
                chart_htmls.append(pio.to_html(fig, include_plotlyjs=False, full_html=False))

            except Exception as e:
                chart_htmls.append(f'<div class="err">⚠️ {ctitle} : {e}</div>')

        # KPIs
        kpi_cards = ""
        for col in num[:5]:
            total = df[col].sum()
            mean  = df[col].mean()
            kpi_cards += f"""
            <div class="kpi">
              <div class="kpi-label">{col}</div>
              <div class="kpi-val">{total:,.0f}</div>
              <div class="kpi-sub">moy. {mean:,.2f}</div>
            </div>"""
        kpi_cards += f"""
            <div class="kpi">
              <div class="kpi-label">Lignes</div>
              <div class="kpi-val">{len(df):,}</div>
              <div class="kpi-sub">{len(df.columns)} colonnes</div>
            </div>"""

        # Stats table
        stats_html = ""
        if num:
            stats = df[num[:6]].describe().round(2)
            rows_html = ""
            for col in stats.columns:
                s = stats[col]
                rows_html += f"""<tr>
                  <td><b>{col}</b></td>
                  <td>{s.get('count','-'):.0f}</td>
                  <td>{s.get('mean',0):.2f}</td>
                  <td>{s.get('std',0):.2f}</td>
                  <td>{s.get('min',0):.2f}</td>
                  <td>{s.get('50%',0):.2f}</td>
                  <td>{s.get('max',0):.2f}</td>
                </tr>"""
            stats_html = f"""
            <table class="tbl">
              <thead><tr><th>Colonne</th><th>N</th><th>Moy.</th><th>Écart-type</th>
                <th>Min</th><th>Médiane</th><th>Max</th></tr></thead>
              <tbody>{rows_html}</tbody>
            </table>"""

        # Preview données
        preview_html = df.head(10).to_html(classes="tbl", border=0, index=False, na_rep="—")

        # Grille graphiques
        n_charts = len(chart_htmls)
        grid = "1fr" if n_charts <= 1 else ("1fr 1fr" if n_charts <= 4 else "1fr 1fr 1fr")
        charts_grid = "\n".join(f'<div class="card chart">{h}</div>' for h in chart_htmls)

        # HTML final
        html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<script src="https://cdn.plot.ly/plotly-2.26.0.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#333}}
.hdr{{background:linear-gradient(135deg,{t['primary']},{t['secondary']});color:{t['text']};
  padding:20px 32px;display:flex;justify-content:space-between;align-items:center}}
.hdr h1{{font-size:22px;font-weight:600}}
.hdr .meta{{font-size:12px;opacity:.8;text-align:right}}
.badge{{background:rgba(255,255,255,.2);padding:2px 10px;border-radius:12px;margin-left:6px;font-size:11px}}
.wrap{{max-width:1500px;margin:0 auto;padding:20px 16px}}
.kpi-row{{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap}}
.kpi{{background:#fff;border-radius:10px;padding:14px 20px;flex:1;min-width:130px;
  box-shadow:0 1px 4px rgba(0,0,0,.08);border-top:3px solid {t['primary']}}}
.kpi-label{{font-size:11px;color:#888;text-transform:uppercase;letter-spacing:.5px}}
.kpi-val{{font-size:26px;font-weight:700;color:{t['primary']};margin:4px 0}}
.kpi-sub{{font-size:11px;color:#aaa}}
.section{{font-size:15px;font-weight:600;color:{t['primary']};margin:20px 0 10px;
  padding-bottom:6px;border-bottom:2px solid {t['accent']}}}
.charts{{display:grid;grid-template-columns:{grid};gap:16px;margin-bottom:20px}}
.card{{background:#fff;border-radius:10px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,.08)}}
.chart .plotly-graph-div{{width:100%!important}}
.tbl{{width:100%;border-collapse:collapse;font-size:12px}}
.tbl th{{background:{t['accent']};color:{t['primary']};padding:8px 10px;text-align:left;
  font-weight:600;border-bottom:2px solid #dde3ee}}
.tbl td{{padding:6px 10px;border-bottom:1px solid #eef0f4}}
.tbl tr:hover td{{background:#f8f9ff}}
.err{{color:#c62828;padding:10px;background:#ffebee;border-radius:6px;font-size:12px}}
.footer{{text-align:center;color:#bbb;font-size:11px;padding:20px 0 30px}}
@media(max-width:768px){{.charts{{grid-template-columns:1fr!important}}}}
</style>
</head>
<body>
<div class="hdr">
  <div>
    <h1>📊 {title}</h1>
    <div style="font-size:12px;opacity:.7;margin-top:4px">
      Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}
    </div>
  </div>
  <div class="meta">
    {f'<div style="font-size:11px">{api_url}</div>' if api_url else ''}
    <span class="badge">{len(df):,} lignes</span>
    <span class="badge">{len(df.columns)} col.</span>
  </div>
</div>
<div class="wrap">
  <div class="kpi-row">{kpi_cards}</div>
  <div class="section">📈 Visualisations</div>
  <div class="charts">{charts_grid}</div>
  {'<div class="section">📐 Statistiques</div><div class="card">' + stats_html + '</div>' if stats_html else ''}
  <div class="section" style="margin-top:20px">📋 Aperçu des données</div>
  <div class="card" style="overflow-x:auto">{preview_html}</div>
</div>
<div class="footer">Agent IA Polyvalent · {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
</body>
</html>"""

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(REPORTS_DIR, f"dashboard_{ts}.html")
        with open(out, "w", encoding="utf-8") as f:
            f.write(html)
        return out

    # ─────────────────────────────────────────────
    # ★ Export PDF professionnel (ReportLab)
    # ─────────────────────────────────────────────

    async def export_pdf(self, data: Any, title: str = "Rapport BI", api_url: str = "") -> str:
        """Rapport PDF professionnel avec ReportLab."""
        return await asyncio.to_thread(self._pdf_sync, data, title, api_url)

    def _pdf_sync(self, data, title, api_url) -> str:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import pandas as pd

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                            Table, TableStyle, Image as RLImage, PageBreak)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
        except ImportError:
            # Fallback : fpdf2
            return self._pdf_fpdf_sync(data, title, api_url)

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        if df is None or df.empty:
            return None

        num = df.select_dtypes(include="number").columns.tolist()
        cat = df.select_dtypes(include="object").columns.tolist()

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_path = os.path.join(REPORTS_DIR, f"rapport_bi_{ts}.pdf")
        tmp_chart = os.path.join(REPORTS_DIR, f"_tmp_chart_{ts}.png")

        # Générer graphique matplotlib
        chart_ok = False
        if cat and num:
            try:
                fig, axes = plt.subplots(1, 2, figsize=(14, 5))
                fig.patch.set_facecolor("white")

                # Graphique 1 : barres
                grp = df.groupby(cat[0])[num[0]].sum()
                if len(grp) <= 20:
                    grp.sort_values(ascending=False).plot(kind="bar", ax=axes[0],
                                                           color="#1a237e", edgecolor="white")
                else:
                    df[num[0]].plot(kind="hist", bins=25, ax=axes[0], color="#1a237e")
                axes[0].set_title(f"{num[0]} par {cat[0]}", fontweight="bold")
                axes[0].tick_params(axis="x", rotation=45)
                axes[0].spines["top"].set_visible(False)
                axes[0].spines["right"].set_visible(False)

                # Graphique 2 : corrélation ou pie
                if len(num) >= 2:
                    corr = df[num[:5]].corr()
                    im = axes[1].imshow(corr, cmap="coolwarm", vmin=-1, vmax=1, aspect="auto")
                    axes[1].set_xticks(range(len(num[:5])))
                    axes[1].set_yticks(range(len(num[:5])))
                    axes[1].set_xticklabels(num[:5], rotation=45, ha="right")
                    axes[1].set_yticklabels(num[:5])
                    axes[1].set_title("Corrélations", fontweight="bold")
                    plt.colorbar(im, ax=axes[1], shrink=0.7)
                else:
                    pie_d = df[cat[0]].value_counts().head(8)
                    axes[1].pie(pie_d.values, labels=pie_d.index, autopct="%1.0f%%",
                                colors=plt.cm.Set3.colors)
                    axes[1].set_title(f"Répartition {cat[0]}", fontweight="bold")

                plt.tight_layout()
                plt.savefig(tmp_chart, dpi=180, bbox_inches="tight", facecolor="white")
                plt.close()
                chart_ok = True
            except Exception as e:
                print(f"[BITool] pdf chart error: {e}")
                plt.close()

        # Construire le PDF avec ReportLab
        doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                                 topMargin=2*cm, bottomMargin=2*cm,
                                 leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        story = []

        # Style titre
        title_style = ParagraphStyle("title", parent=styles["Title"],
                                      fontSize=20, spaceAfter=6, textColor=colors.HexColor("#1a237e"))
        h2_style = ParagraphStyle("h2", parent=styles["Heading2"],
                                   fontSize=13, spaceBefore=12, spaceAfter=6,
                                   textColor=colors.HexColor("#0d47a1"))
        body_style = ParagraphStyle("body", parent=styles["Normal"],
                                     fontSize=9, spaceAfter=3)

        # En-tête
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", body_style))
        if api_url:
            story.append(Paragraph(f"Source : {api_url}", body_style))
        story.append(Spacer(1, 0.4*cm))

        # KPIs table
        story.append(Paragraph("Indicateurs clés", h2_style))
        kpi_data = [["Indicateur", "Total", "Moyenne", "Min", "Max"]]
        for col in num[:6]:
            kpi_data.append([
                col,
                f"{df[col].sum():,.2f}",
                f"{df[col].mean():,.2f}",
                f"{df[col].min():,.2f}",
                f"{df[col].max():,.2f}",
            ])
        kpi_data.append(["Lignes totales", f"{len(df):,}", "—", "—", "—"])

        kpi_table = Table(kpi_data, colWidths=[5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f5f5ff"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#ddddee")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.5*cm))

        # Graphique
        if chart_ok and Path(tmp_chart).exists():
            story.append(Paragraph("Visualisations", h2_style))
            img = RLImage(tmp_chart, width=17*cm, height=7*cm)
            story.append(img)
            story.append(Spacer(1, 0.4*cm))

        # Statistiques descriptives
        if num:
            story.append(PageBreak())
            story.append(Paragraph("Statistiques descriptives", h2_style))
            stats = df[num[:8]].describe().round(2)
            stat_data = [[""] + list(stats.columns)]
            for idx in stats.index:
                stat_data.append([idx] + [str(v) for v in stats.loc[idx]])
            st = Table(stat_data)
            st.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8eaf6")),
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8eaf6")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ccccdd")),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(st)

        # Aperçu données
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("Aperçu des données (30 premières lignes)", h2_style))
        disp_cols = df.columns[:8].tolist()
        rows_data = [disp_cols]
        for _, row in df.head(30).iterrows():
            rows_data.append([str(row[c])[:18] for c in disp_cols])
        col_w = 17*cm / len(disp_cols)
        data_table = Table(rows_data, colWidths=[col_w] * len(disp_cols))
        data_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9ff")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#ddddee")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(data_table)

        doc.build(story)

        # Nettoyage
        try:
            os.unlink(tmp_chart)
        except Exception:
            pass

        return pdf_path

    def _pdf_fpdf_sync(self, data, title, api_url) -> str:
        """Fallback PDF avec fpdf2."""
        from fpdf import FPDF
        import pandas as pd

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(REPORTS_DIR, f"rapport_bi_{ts}.pdf")
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, title, ln=True)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 5, f"Généré le {datetime.now().strftime('%d/%m/%Y')}", ln=True)
        if api_url:
            pdf.cell(0, 5, f"Source : {api_url}", ln=True)
        pdf.ln(3)
        if df is not None:
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(0, 6, f"Données : {len(df)} lignes x {len(df.columns)} colonnes", ln=True)
            pdf.set_font("Helvetica", "", 8)
            for line in df.head(20).to_string(index=False).split("\n"):
                try:
                    pdf.cell(0, 4, line[:100], ln=True)
                except Exception:
                    pass
        pdf.output(path)
        return path

    # ─────────────────────────────────────────────
    # Export Excel (bonus)
    # ─────────────────────────────────────────────

    async def export_excel(self, data: Any, title: str = "Rapport", sheets: dict = None) -> str:
        """Exporter vers Excel avec mise en forme (openpyxl)."""
        return await asyncio.to_thread(self._excel_sync, data, title, sheets)

    def _excel_sync(self, data, title, sheets) -> str:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        df = data if isinstance(data, pd.DataFrame) else self.normalize_data(data)
        if df is None:
            return None

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(REPORTS_DIR, f"rapport_{ts}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = title[:30]

        # Style header
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill(fill_type="solid", fgColor="1A237E")
        alt_fill     = PatternFill(fill_type="solid", fgColor="E8EAF6")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border  = Border(
            bottom=Side(style="thin", color="CCCCDD"),
            right=Side(style="thin", color="CCCCDD"),
        )

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                elif r_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.border = thin_border

        # Ajuster largeur colonnes
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Feuille statistiques
        num = df.select_dtypes(include="number").columns.tolist()
        if num:
            ws2 = wb.create_sheet("Statistiques")
            stats = df[num].describe().round(2).reset_index()
            for r_idx, row in enumerate(dataframe_to_rows(stats, index=False, header=True), 1):
                for c_idx, val in enumerate(row, 1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill

        wb.save(path)
        return path

    # ─────────────────────────────────────────────
    # Pipeline complet
    # ─────────────────────────────────────────────

    async def full_report(
        self,
        api_url: str = None,
        api_key: str = None,
        auth_type: str = "bearer",
        title: str = "Rapport BI",
        chart_configs: list[dict] = None,
        export: str = "html",   # html | pdf | excel | all
        data: Any = None,
        theme: str = "blue",
    ) -> dict:
        """Pipeline : données → dashboard HTML + PDF + Excel."""
        import pandas as pd

        # Récupérer données
        if data is None and api_url:
            r = await self.fetch_api_data(url=api_url, api_key=api_key, auth_type=auth_type)
            if not r["success"]:
                return {"success": False, "error": r["error"]}
            raw = r["data"]
        elif data is not None:
            raw = data
        else:
            return {"success": False, "error": "Fournir api_url ou data."}

        df = raw if isinstance(raw, pd.DataFrame) else self.normalize_data(raw)
        if df is None or df.empty:
            return {"success": False, "error": "Données vides ou non convertibles en tableau."}

        result = {
            "success": True,
            "shape": f"{len(df)} lignes × {len(df.columns)} colonnes",
            "columns": df.columns.tolist(),
            "html_path": None, "pdf_path": None, "excel_path": None,
        }

        if export in ("html", "all"):
            result["html_path"] = await self.generate_dashboard(
                df, title, chart_configs, api_url or "", theme
            )
        if export in ("pdf", "all"):
            result["pdf_path"] = await self.export_pdf(df, title, api_url or "")
        if export in ("excel", "all"):
            result["excel_path"] = await self.export_excel(df, title)

        return result

    # ─────────────────────────────────────────────
    # Planification APScheduler
    # ─────────────────────────────────────────────

    def schedule_report(
        self,
        api_url: str,
        title: str,
        cron: str = None,
        export: str = "html",
        api_key: str = None,
        on_ready=None,   # async callback(path)
    ) -> dict:
        """Planifier un rapport récurrent (expression cron)."""
        try:
            from apscheduler.schedulers.asyncio import AsyncIOScheduler
            from apscheduler.triggers.cron import CronTrigger

            if not self._scheduler:
                self._scheduler = AsyncIOScheduler()

            expr = cron or os.getenv("BI_REPORT_SCHEDULE", "0 8 * * 1-5")
            parts = (expr.split() + ["*"] * 5)[:5]
            minute, hour, day, month, dow = parts

            async def _job():
                r = await self.full_report(
                    api_url=api_url, api_key=api_key,
                    title=f"{title} — {datetime.now().strftime('%d/%m/%Y')}",
                    export=export,
                )
                if on_ready and r.get("html_path"):
                    await on_ready(r.get("html_path") or r.get("pdf_path"))

            job_id = f"bi_{hashlib.md5(api_url.encode()).hexdigest()[:8]}"
            self._scheduler.add_job(
                _job,
                CronTrigger(minute=minute, hour=hour, day=day, month=month, day_of_week=dow),
                id=job_id, replace_existing=True, name=f"BI: {title}",
            )
            if not self._scheduler.running:
                self._scheduler.start()

            return {"success": True, "message": f"Rapport '{title}' planifié : {expr}", "id": job_id}

        except ImportError:
            return {"success": False, "message": "apscheduler non installé."}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def list_scheduled(self) -> list[dict]:
        if not self._scheduler:
            return []
        return [{"id": j.id, "name": j.name, "next_run": str(j.next_run_time)}
                for j in self._scheduler.get_jobs()]

    def cancel_schedule(self, job_id: str) -> bool:
        if self._scheduler:
            try:
                self._scheduler.remove_job(job_id)
                return True
            except Exception:
                pass
        return False
